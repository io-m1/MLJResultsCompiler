#!/usr/bin/env python3
"""
Obstetrics & Gynecology Test Results Collation Automation
Processes individual test result sheets and compiles into unified result sheet
Designed for monthly automated processing with error tracking
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from pathlib import Path
from datetime import datetime
import json
import sys
from typing import Dict, List, Tuple

class TestResultsCollator:
    def __init__(self, input_dir: str, output_dir: str, month_year: str):
        self.input_dir = input_dir
        self.output_dir = output_dir
        self.month_year = month_year
        self.test_files = []
        self.error_log = {
            'errors': [],
            'warnings': [],
            'processed_files': [],
            'timestamp': datetime.now().isoformat()
        }
        self.pass_mark = 50  # Default pass mark percentage
        
    def discover_test_files(self) -> Dict[int, str]:
        """Discover test files in input directory and map to test numbers"""
        test_mapping = {}
        patterns = {
            1: ['TEST_1', 'test_1'],
            2: ['TEST_2', 'test_2'],
            3: ['TEST_3', 'test_3'],
            4: ['TEST_4', 'test_4'],
            5: ['TEST_5', 'test_5', 'Ultrasonography_Test_5']
        }
        
        for file in os.listdir(self.input_dir):
            if not file.endswith('.xlsx'):
                continue
                
            for test_num, patterns_list in patterns.items():
                if any(pattern in file for pattern in patterns_list):
                    test_mapping[test_num] = os.path.join(self.input_dir, file)
                    self.error_log['processed_files'].append({
                        'test': f'TEST_{test_num}',
                        'filename': file
                    })
                    break
        
        if not test_mapping:
            raise FileNotFoundError(f"No test files found in {self.input_dir}")
            
        return test_mapping
    
    def load_test_sheet(self, filepath: str) -> pd.DataFrame:
        """Load a test sheet and extract relevant columns"""
        try:
            # Read the Responses sheet
            df = pd.read_excel(filepath, sheet_name='Responses')
            
            # Standardize column names (handle variations in naming)
            col_mapping = {
                ' Full Names': 'Full Names',
                'Full names': 'Full Names',
                'Full Names': 'Full Names',
                'Email': 'Email',
                'email': 'Email',
                'Result': 'Result',
                'result': 'Result'
            }
            
            for old_col, new_col in col_mapping.items():
                if old_col in df.columns:
                    df.rename(columns={old_col: new_col}, inplace=True)
            
            # Extract only needed columns
            required_cols = ['Full Names', 'Email', 'Result']
            available_cols = [col for col in required_cols if col in df.columns]
            
            if not available_cols:
                raise ValueError(f"Required columns not found. Available: {df.columns.tolist()}")
            
            df_subset = df[available_cols].copy()
            
            # Clean data
            df_subset['Full Names'] = df_subset['Full Names'].str.strip()
            df_subset['Email'] = df_subset['Email'].str.strip()
            
            # Convert Result to percentage (remove % and convert)
            df_subset['Result'] = df_subset['Result'].astype(str).str.rstrip('%')
            
            return df_subset
            
        except Exception as e:
            self.error_log['errors'].append({
                'file': filepath,
                'error': f"Failed to load test sheet: {str(e)}"
            })
            return pd.DataFrame()
    
    def merge_test_results(self, test_mapping: Dict[int, str]) -> pd.DataFrame:
        """Merge all test results into single dataframe with participant as key"""
        master_df = None
        
        for test_num in sorted(test_mapping.keys()):
            filepath = test_mapping[test_num]
            
            if not os.path.exists(filepath):
                self.error_log['errors'].append({
                    'test': f'TEST_{test_num}',
                    'error': f'File not found: {filepath}'
                })
                continue
            
            test_df = self.load_test_sheet(filepath)
            
            if test_df.empty:
                self.error_log['warnings'].append({
                    'test': f'TEST_{test_num}',
                    'warning': 'Empty or invalid data extracted'
                })
                continue
            
            # Create merge key (case-insensitive, strip whitespace)
            test_df['merge_key'] = test_df['Full Names'].str.lower().str.strip()
            test_df = test_df.rename(columns={'Result': f'TEST_{test_num}'})
            
            if master_df is None:
                master_df = test_df[['Full Names', 'Email', 'merge_key', f'TEST_{test_num}']].copy()
            else:
                # Merge on key columns
                merge_cols = ['merge_key', f'TEST_{test_num}']
                master_df = master_df.merge(
                    test_df[['merge_key', f'TEST_{test_num}']],
                    on='merge_key',
                    how='outer'
                )
                
                # Update participant name/email if missing
                test_df_name = test_df[['merge_key', 'Full Names', 'Email']].copy()
                test_df_name = test_df_name.drop_duplicates(subset='merge_key')
                
                for col in ['Full Names', 'Email']:
                    mask = master_df[col].isna()
                    if mask.any():
                        merged = master_df[mask][['merge_key']].merge(
                            test_df_name[['merge_key', col]],
                            on='merge_key',
                            how='left'
                        )
                        master_df.loc[mask, col] = merged[col].values
        
        if master_df is None or master_df.empty:
            raise ValueError("No test data could be merged")
        
        # Clean up
        master_df = master_df.drop_duplicates(subset='merge_key')
        master_df = master_df.drop(columns=['merge_key'])
        master_df = master_df.sort_values('Full Names').reset_index(drop=True)
        
        return master_df
    
    def create_final_sheet(self, master_df: pd.DataFrame) -> openpyxl.Workbook:
        """Create the final result sheet with formatting"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Responses'
        
        # Header row
        headers = ['S/N', 'NAMES', 'EMAIL', 'TEST 1', 'TEST 2', 'TEST 3', 
                   'TEST 4', 'TEST 5', 'GRP DISCUSSION', 'TOTAL MARK', 'SCORE', 'STATUS']
        
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Data rows
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for idx, row in master_df.iterrows():
            row_num = idx + 2
            ws[f'A{row_num}'] = idx + 1  # S/N
            ws[f'B{row_num}'] = row['Full Names']
            ws[f'C{row_num}'] = row['Email']
            
            # Test scores (columns D-H)
            test_cols = ['TEST_1', 'TEST_2', 'TEST_3', 'TEST_4', 'TEST_5']
            for col_idx, test_col in enumerate(test_cols):
                col_letter = get_column_letter(4 + col_idx)
                
                if test_col in row and pd.notna(row[test_col]):
                    try:
                        val = float(row[test_col])
                        # Normalize to percentage if < 100
                        if val <= 100:
                            ws[f'{col_letter}{row_num}'] = val / 100
                        else:
                            ws[f'{col_letter}{row_num}'] = val
                        ws[f'{col_letter}{row_num}'].number_format = '0.0%'
                    except:
                        ws[f'{col_letter}{row_num}'] = None
                else:
                    ws[f'{col_letter}{row_num}'] = None
            
            # GRP DISCUSSION (fixed value)
            ws[f'I{row_num}'] = 0.8
            
            # TOTAL MARK formula (sum of D-I)
            ws[f'J{row_num}'] = f'=SUM(D{row_num}:I{row_num})'
            
            # SCORE formula (TOTAL * weight factor)
            ws[f'K{row_num}'] = f'=J{row_num}*16.6666'
            
            # STATUS formula (IF score >= 50 PASS else FAIL)
            ws[f'L{row_num}'] = f'=IF(K{row_num}>={self.pass_mark},"PASS","FAIL")'
            
            # Apply formatting to data rows
            for col in range(1, 13):
                cell = ws.cell(row=row_num, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Percentage formatting for test columns
                if col >= 4 and col <= 9:
                    if cell.number_format != '0.0%':
                        cell.number_format = '0.0%'
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            ws.column_dimensions[col].width = 12
        
        # Freeze panes
        ws.freeze_panes = 'A2'
        
        return wb
    
    def save_results(self, wb: openpyxl.Workbook, filename: str = None):
        """Save workbook to output directory"""
        if filename is None:
            filename = f"OBS_{self.month_year}_RESULT_SHEET.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        os.makedirs(self.output_dir, exist_ok=True)
        
        wb.save(filepath)
        self.error_log['output_file'] = filepath
        return filepath
    
    def save_error_log(self):
        """Save error log as JSON"""
        log_filename = f"collation_log_{self.month_year}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        log_path = os.path.join(self.output_dir, log_filename)
        
        os.makedirs(self.output_dir, exist_ok=True)
        
        with open(log_path, 'w') as f:
            json.dump(self.error_log, f, indent=2)
        
        return log_path
    
    def run(self) -> Tuple[str, bool]:
        """Execute the full collation process"""
        try:
            print(f"[INFO] Starting test results collation for {self.month_year}")
            
            # Step 1: Discover test files
            print("[STEP 1] Discovering test files...")
            test_mapping = self.discover_test_files()
            print(f"  Found tests: {sorted(test_mapping.keys())}")
            
            # Step 2: Merge test results
            print("[STEP 2] Merging test results...")
            master_df = self.merge_test_results(test_mapping)
            print(f"  Merged {len(master_df)} participants")
            
            # Step 3: Create final sheet
            print("[STEP 3] Creating final result sheet...")
            wb = self.create_final_sheet(master_df)
            
            # Step 4: Save results
            print("[STEP 4] Saving files...")
            output_path = self.save_results(wb)
            log_path = self.save_error_log()
            
            print(f"  Output: {output_path}")
            print(f"  Log: {log_path}")
            
            success = len(self.error_log['errors']) == 0
            
            if success:
                print("[SUCCESS] Collation completed without errors")
            else:
                print(f"[WARNING] Collation completed with {len(self.error_log['errors'])} errors")
            
            return output_path, success
            
        except Exception as e:
            self.error_log['errors'].append({
                'error': f"Fatal error: {str(e)}"
            })
            self.save_error_log()
            print(f"[ERROR] {str(e)}")
            return None, False


def main():
    if len(sys.argv) < 3:
        print("Usage: python test_collation_automation.py <input_dir> <output_dir> [month_year]")
        print("Example: python test_collation_automation.py ./input ./output JAN_2026")
        sys.exit(1)
    
    input_dir = sys.argv[1]
    output_dir = sys.argv[2]
    month_year = sys.argv[3] if len(sys.argv) > 3 else datetime.now().strftime('%b_%Y').upper()
    
    collator = TestResultsCollator(input_dir, output_dir, month_year)
    output_path, success = collator.run()
    
    sys.exit(0 if success else 1)


if __name__ == '__main__':
    main()
