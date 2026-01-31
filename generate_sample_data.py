"""Generate sample test data files for validation"""
import openpyxl
from pathlib import Path

# Ensure input directory exists
input_dir = Path("input")
input_dir.mkdir(exist_ok=True)

# Sample participant data
participants = [
    ("Alice Johnson", "alice.johnson@email.com", 92),
    ("Bob Smith", "bob.smith@email.com", 88),
    ("Charlie Brown", "charlie.brown@email.com", 95),
    ("Diana Prince", "diana.prince@email.com", 87),
    ("Emma Wilson", "emma.wilson@email.com", 91),
]

# Test 1 - All participants
wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.title = "Results"
ws1['A1'] = "Full Name"
ws1['B1'] = "Email"
ws1['C1'] = "Score"

for i, (name, email, score) in enumerate(participants, start=2):
    ws1[f'A{i}'] = name
    ws1[f'B{i}'] = email
    ws1[f'C{i}'] = score

ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 25
ws1.column_dimensions['C'].width = 12

wb1.save(input_dir / 'Test 1.xlsx')
print("✓ Test 1.xlsx created (5 participants)")

# Test 2 - 4 participants (missing Emma)
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "Results"
ws2['A1'] = "Full Name"
ws2['B1'] = "Email"
ws2['C1'] = "Score %"

test2_data = [
    ("Alice Johnson", "alice.johnson@email.com", 89),
    ("Bob Smith", "bob.smith@email.com", 91),
    ("Charlie Brown", "charlie.brown@email.com", 93),
    ("Diana Prince", "diana.prince@email.com", 85),
]

for i, (name, email, score) in enumerate(test2_data, start=2):
    ws2[f'A{i}'] = name
    ws2[f'B{i}'] = email
    ws2[f'C{i}'] = f"{score}%"

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 25
ws2.column_dimensions['C'].width = 12

wb2.save(input_dir / 'Test 2.xlsx')
print("✓ Test 2.xlsx created (4 participants)")

# Test 3 - 3 participants (different subset)
wb3 = openpyxl.Workbook()
ws3 = wb3.active
ws3.title = "Results"
ws3['A1'] = "Participant Name"
ws3['B1'] = "E-mail"
ws3['C1'] = "Result"

test3_data = [
    ("Alice Johnson", "alice.johnson@email.com", 94),
    ("Charlie Brown", "charlie.brown@email.com", 96),
    ("Emma Wilson", "emma.wilson@email.com", 88),
]

for i, (name, email, score) in enumerate(test3_data, start=2):
    ws3[f'A{i}'] = name
    ws3[f'B{i}'] = email
    ws3[f'C{i}'] = score

ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 25
ws3.column_dimensions['C'].width = 12

wb3.save(input_dir / 'Test 3.xlsx')
print("✓ Test 3.xlsx created (3 participants)")

print("\nSample data ready in input/ folder!")
