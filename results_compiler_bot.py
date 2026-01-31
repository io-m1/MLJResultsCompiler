#!/usr/bin/env python3
"""
MLJResultsCompiler - Results Compilation Automation Bot
Consolidates 5 test result sheets from SurveyHeart into one comprehensive report
Author: MLJ Results Compiler
Date: January 31, 2026
Version: 1.0 - Production Ready
"""

import os
import sys
import logging
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Tuple, List, Optional, Dict
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import glob

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('compiler_execution.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class ResultsCompiler:
    """Main class for compiling test results from multiple test files"""
    
    def __init__(self, input_folder: str = 'input', output_folder: str = 'output'):
        """Initialize the compiler with input and output folders"""
        self.input_folder = Path(input_folder)
        self.output_folder = Path(output_folder)
        self.test_files = {}
        self.dataframes = {}
        self.merged_df = None
        self.final_df = None
        
        # Create output folder if it doesn't exist
        self.output_folder.mkdir(exist_ok=True)
        
        # Color mapping for each test
        self.colors = {
            'Test 1': 'FFFFFF',  # White
            'Test 2': '87CEEB',  # Sky Blue
            'Test 3': 'FFFF00',  # Yellow
            'Test 4': '556B2F',  # Army Green
            'Test 5': 'FF0000',  # Red
        }
        
        logger.info(f"ResultsCompiler initialized")
        logger.info(f"Input folder: {self.input_folder.absolute()}")
        logger.info(f"Output folder: {self.output_folder.absolute()}")
    
    def find_test_files(self) -> bool:
        """Find all 5 test files in the input folder"""
        logger.info("Searching for test files...")
        
        if not self.input_folder.exists():
            logger.error(f"Input folder not found: {self.input_folder}")
            return False
        
        test_patterns = {
            'Test 1': ['TEST_1*', 'test_1*', 'Test_1*'],
            'Test 2': ['TEST_2*', 'test_2*', 'Test_2*'],
            'Test 3': ['TEST_3*', 'test_3*', 'Test_3*'],
            'Test 4': ['TEST_4*', 'test_4*', 'Test_4*'],
            'Test 5': ['TEST_5*', 'test_5*', 'Test_5*'],
        }
        
        for test_name, patterns in test_patterns.items():
            found = False
            for pattern in patterns:
                matches = list(self.input_folder.glob(f"{pattern}.xlsx"))
                if matches:
                    self.test_files[test_name] = matches[0]
                    logger.info(f"Found {test_name}: {matches[0].name}")
                    found = True
                    break
            
            if not found:
                logger.warning(f"Could not find {test_name} file")
        
        if len(self.test_files) < 5:
            logger.warning(f"Only found {len(self.test_files)} of 5 test files")
        else:
            logger.info(f"✓ All 5 test files found successfully")
        
        return len(self.test_files) >= 5
    
    def detect_column_names(self, df: pd.DataFrame) -> Tuple[Optional[str], Optional[str], Optional[str]]:
        """
        Detect column names for Full Name, Email, and Score/Result
        Handles variations like: "Full Names", "Full names", "Name", etc.
        """
        columns = df.columns.tolist()
        
        # Detect Name column (try multiple variations)
        name_col = None
        name_variations = [col for col in columns if col.strip().lower() in 
                          ['full name', 'full names', 'name', 'participant', 'participant name']]
        if name_variations:
            name_col = name_variations[0]
        
        # Detect Email column
        email_col = None
        email_variations = [col for col in columns if col.strip().lower() in 
                           ['email', 'e-mail', 'email address', 'e_mail']]
        if email_variations:
            email_col = email_variations[0]
        
        # Detect Score/Result column
        score_col = None
        score_variations = [col for col in columns if col.strip().lower() in 
                           ['score', 'result', 'percentage', 'marks', 'points']]
        if score_variations:
            score_col = score_variations[0]
        
        logger.debug(f"Detected columns: Name={name_col}, Email={email_col}, Score={score_col}")
        return name_col, email_col, score_col
    
    def load_test_file(self, test_name: str, filepath: Path) -> bool:
        """Load and extract data from a single test file"""
        logger.info(f"Loading {test_name} from {filepath.name}...")
        
        try:
            # Read XLSX file
            df = pd.read_excel(filepath)
            logger.debug(f"{test_name}: Loaded {len(df)} rows, {len(df.columns)} columns")
            
            # Detect column names
            name_col, email_col, score_col = self.detect_column_names(df)
            
            if not all([name_col, email_col, score_col]):
                logger.error(f"{test_name}: Could not detect all required columns")
                logger.error(f"  Name column: {name_col}, Email: {email_col}, Score: {score_col}")
                return False
            
            # Extract only required columns
            df_extracted = df[[name_col, email_col, score_col]].copy()
            
            # Rename columns to standard names
            df_extracted.columns = ['Full Name', 'Email', f'{test_name} Score']
            
            # Clean data
            df_extracted['Full Name'] = df_extracted['Full Name'].astype(str).str.strip()
            df_extracted['Email'] = df_extracted['Email'].astype(str).str.strip().str.lower()
            
            # Parse score - remove % and convert to float
            def parse_score(score):
                if pd.isna(score) or score == '':
                    return np.nan
                score_str = str(score).strip().replace('%', '')
                try:
                    return float(score_str)
                except ValueError:
                    return np.nan
            
            df_extracted[f'{test_name} Score'] = df_extracted[f'{test_name} Score'].apply(parse_score)
            
            # Remove rows with missing name or email
            df_extracted = df_extracted.dropna(subset=['Full Name', 'Email'])
            df_extracted = df_extracted[df_extracted['Email'] != '']
            
            logger.info(f"✓ {test_name}: Extracted {len(df_extracted)} participants")
            self.dataframes[test_name] = df_extracted
            return True
            
        except Exception as e:
            logger.error(f"Error loading {test_name}: {str(e)}")
            return False
    
    def load_all_test_files(self) -> bool:
        """Load all test files"""
        logger.info("Loading all test files...")
        
        success_count = 0
        for test_name in ['Test 1', 'Test 2', 'Test 3', 'Test 4', 'Test 5']:
            if test_name in self.test_files:
                if self.load_test_file(test_name, self.test_files[test_name]):
                    success_count += 1
        
        logger.info(f"Successfully loaded {success_count} of {len(self.test_files)} test files")
        return success_count == len(self.test_files)
    
    def merge_tests(self) -> bool:
        """Merge all test dataframes on email (outer join)"""
        logger.info("Merging test results on email...")
        
        if not self.dataframes:
            logger.error("No test dataframes loaded")
            return False
        
        try:
            # Start with Test 1 as base
            if 'Test 1' not in self.dataframes:
                logger.error("Test 1 not found - required as base")
                return False
            
            self.merged_df = self.dataframes['Test 1'].copy()
            logger.debug(f"Starting merge with Test 1: {len(self.merged_df)} rows")
            
            # Merge with Tests 2-5 (outer join)
            for test_num in [2, 3, 4, 5]:
                test_name = f'Test {test_num}'
                if test_name in self.dataframes:
                    # Merge on email
                    self.merged_df = pd.merge(
                        self.merged_df,
                        self.dataframes[test_name],
                        on='Email',
                        how='outer',
                        suffixes=('', f'_{test_num}')
                    )
                    
                    # Handle name column conflicts - keep from Test 1
                    if f'Full Name_{test_num}' in self.merged_df.columns:
                        self.merged_df['Full Name'] = self.merged_df['Full Name'].fillna(
                            self.merged_df[f'Full Name_{test_num}']
                        )
                        self.merged_df.drop(f'Full Name_{test_num}', axis=1, inplace=True)
                    
                    logger.debug(f"After merging {test_name}: {len(self.merged_df)} rows")
            
            logger.info(f"✓ Merge complete: {len(self.merged_df)} unique participants")
            return True
            
        except Exception as e:
            logger.error(f"Error during merge: {str(e)}")
            return False
    
    def clean_and_sort(self) -> bool:
        """Remove duplicates and sort alphabetically"""
        logger.info("Cleaning and sorting results...")
        
        try:
            # Remove duplicate emails (keep first)
            initial_count = len(self.merged_df)
            self.merged_df = self.merged_df.drop_duplicates(subset=['Email'], keep='first')
            duplicates_removed = initial_count - len(self.merged_df)
            
            if duplicates_removed > 0:
                logger.info(f"Removed {duplicates_removed} duplicate emails")
            
            # Sort alphabetically by Full Name (case-insensitive)
            self.merged_df = self.merged_df.sort_values(
                'Full Name',
                key=lambda x: x.str.lower(),
                ignore_index=True
            )
            
            logger.info(f"✓ Sorted {len(self.merged_df)} participants alphabetically")
            return True
            
        except Exception as e:
            logger.error(f"Error during cleanup: {str(e)}")
            return False
    
    def format_scores(self) -> bool:
        """Format scores as percentages with % sign"""
        logger.info("Formatting scores...")
        
        try:
            score_columns = [col for col in self.merged_df.columns if 'Score' in col]
            
            for col in score_columns:
                # Format as percentage string (keep 1 decimal place)
                self.merged_df[col] = self.merged_df[col].apply(
                    lambda x: f"{x:.1f}%" if pd.notna(x) else ""
                )
            
            logger.info(f"✓ Formatted {len(score_columns)} score columns")
            return True
            
        except Exception as e:
            logger.error(f"Error formatting scores: {str(e)}")
            return False
    
    def export_to_xlsx(self, filename: str = 'Consolidated_Results.xlsx') -> bool:
        """Export merged dataframe to formatted XLSX file"""
        logger.info(f"Exporting to {filename}...")
        
        try:
            output_path = self.output_folder / filename
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Results'
            
            # Define styles
            header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            header_font = Font(bold=True, size=12)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write header row
            headers = self.merged_df.columns.tolist()
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Write data rows with colors
            for row_idx, (_, row) in enumerate(self.merged_df.iterrows(), 2):
                for col_idx, (col_name, value) in enumerate(zip(headers, row), 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    
                    # Apply color based on test column
                    if 'Test 1' in col_name:
                        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                    elif 'Test 2' in col_name:
                        cell.fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')
                    elif 'Test 3' in col_name:
                        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    elif 'Test 4' in col_name:
                        cell.fill = PatternFill(start_color='556B2F', end_color='556B2F', fill_type='solid')
                    elif 'Test 5' in col_name:
                        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    
                    # Center alignment for scores
                    if 'Score' in col_name:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for col_idx, header in enumerate(headers, 1):
                max_length = len(str(header))
                for row in self.merged_df.itertuples(index=False):
                    max_length = max(max_length, len(str(row[col_idx-1])))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"✓ Exported to {output_path}")
            logger.info(f"✓ File contains {len(self.merged_df)} participants")
            
            return True
            
        except Exception as e:
            logger.error(f"Error exporting to XLSX: {str(e)}")
            return False
    
    def generate_report(self):
        """Generate and print summary report"""
        logger.info("\n" + "="*60)
        logger.info("COMPILATION SUMMARY")
        logger.info("="*60)
        
        if self.merged_df is not None:
            logger.info(f"Total Unique Participants: {len(self.merged_df)}")
            
            # Count participants by test
            for test_num in [1, 2, 3, 4, 5]:
                score_col = f'Test {test_num} Score'
                if score_col in self.merged_df.columns:
                    non_empty = (self.merged_df[score_col] != '').sum()
                    logger.info(f"Test {test_num}: {non_empty} participants")
            
            logger.info(f"\nOutput file: {self.output_folder / 'Consolidated_Results.xlsx'}")
            logger.info(f"Log file: compiler_execution.log")
            logger.info("="*60 + "\n")
    
    def run(self) -> bool:
        """Execute the full compilation pipeline"""
        logger.info("Starting MLJ Results Compilation Bot...")
        logger.info(f"Timestamp: {datetime.now()}")
        
        start_time = datetime.now()
        
        # Step 1: Find test files
        if not self.find_test_files():
            logger.error("Could not find all required test files")
            return False
        
        # Step 2: Load test files
        if not self.load_all_test_files():
            logger.error("Could not load all test files")
            return False
        
        # Step 3: Merge tests
        if not self.merge_tests():
            logger.error("Could not merge test results")
            return False
        
        # Step 4: Clean and sort
        if not self.clean_and_sort():
            logger.error("Could not clean and sort results")
            return False
        
        # Step 5: Format scores
        if not self.format_scores():
            logger.error("Could not format scores")
            return False
        
        # Step 6: Export to XLSX
        if not self.export_to_xlsx():
            logger.error("Could not export results")
            return False
        
        # Generate report
        self.generate_report()
        
        end_time = datetime.now()
        duration = (end_time - start_time).total_seconds()
        
        logger.info(f"✓ COMPILATION COMPLETE in {duration:.2f} seconds")
        return True


def main():
    """Main entry point"""
    # Parse command line arguments
    input_folder = sys.argv[1] if len(sys.argv) > 1 else 'input'
    output_folder = sys.argv[2] if len(sys.argv) > 2 else 'output'
    
    # Create compiler and run
    compiler = ResultsCompiler(input_folder=input_folder, output_folder=output_folder)
    
    if compiler.run():
        sys.exit(0)
    else:
        sys.exit(1)


if __name__ == '__main__':
    main()
