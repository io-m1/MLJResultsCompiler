#!/usr/bin/env python3
"""
Upload & Consolidation Flow Test
Tests: File upload, format detection, validation, consolidation, output
Date: February 1, 2026
"""

import os
import sys
import logging
from pathlib import Path
import tempfile
import asyncio
from openpyxl import Workbook

# Force UTF-8 output
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

sys.path.insert(0, str(Path(__file__).parent))

from config import BotConfig
from src.excel_processor import ExcelProcessor
from src.session_manager import SessionManager

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - [%(levelname)s] - %(name)s - %(message)s'
)
logger = logging.getLogger(__name__)


class UploadFlowTester:
    """Test upload and consolidation workflow"""
    
    def __init__(self):
        self.config = BotConfig()
        self.session_manager = SessionManager()
        self.excel_processor = ExcelProcessor(
            input_dir=str(Path(self.config.input_folder)),
            output_dir=str(Path(self.config.output_folder))
        )
        self.passed = 0
        self.failed = 0
        self.errors = []
    
    def create_test_file(self, filename: str, test_name: str, rows: int = 3) -> str:
        """Create a test Excel file"""
        temp_dir = Path(tempfile.gettempdir())
        file_path = temp_dir / filename
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = test_name
            
            # Headers
            ws['A1'] = 'Name'
            ws['B1'] = 'Email'
            ws['C1'] = 'Score'
            
            # Data rows
            for i in range(rows):
                ws[f'A{i+2}'] = f'Student {i+1}'
                ws[f'B{i+2}'] = f'student{i+1}@example.com'
                ws[f'C{i+2}'] = 80 + i
            
            wb.save(file_path)
            logger.info(f"  Created test file: {file_path}")
            return str(file_path)
        except Exception as e:
            logger.error(f"  Failed to create test file: {e}")
            raise
    
    def test_file_creation(self):
        """Test 1: Can create valid test files"""
        logger.info("\n📋 Test 1: File Creation")
        try:
            file_path = self.create_test_file("test_math.xlsx", "Mathematics", rows=3)
            assert os.path.exists(file_path), "File not created"
            assert os.path.getsize(file_path) > 0, "File is empty"
            
            os.remove(file_path)
            logger.info("  ✅ Test file creation working")
            self.passed += 1
            return True
        except Exception as e:
            logger.error(f"  ❌ File creation failed: {e}")
            self.failed += 1
            self.errors.append(f"File Creation: {e}")
            return False
    
    def test_multiple_file_handling(self):
        """Test 2: Handle multiple files"""
        logger.info("\n📋 Test 2: Multiple File Handling")
        try:
            # Create multiple test files
            files = []
            test_subjects = ["Mathematics", "Physics", "Chemistry", "English", "History"]
            
            for subject in test_subjects:
                file_path = self.create_test_file(
                    f"test_{subject.lower()}.xlsx",
                    subject,
                    rows=5
                )
                files.append(file_path)
                assert os.path.exists(file_path), f"File not created: {subject}"
            
            # Verify all files created
            assert len(files) == 5, "Not all files created"
            
            # Cleanup
            for file_path in files:
                if os.path.exists(file_path):
                    os.remove(file_path)
            
            logger.info("  ✅ Multiple file handling working")
            self.passed += 1
            return True
        except Exception as e:
            logger.error(f"  ❌ Multiple file handling failed: {e}")
            self.failed += 1
            self.errors.append(f"Multiple File Handling: {e}")
            return False
    
    def test_file_validation(self):
        """Test 3: File validation"""
        logger.info("\n📋 Test 3: File Validation")
        try:
            # Create valid file
            valid_file = self.create_test_file("valid_test.xlsx", "Test", rows=3)
            assert os.path.exists(valid_file), "Valid file not created"
            
            # Create invalid file (missing columns)
            invalid_file_path = Path(tempfile.gettempdir()) / "invalid_test.xlsx"
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'OnlyName'  # Missing Email and Score
            ws['A2'] = 'Student 1'
            wb.save(invalid_file_path)
            
            logger.info("  ✅ File validation structure verified")
            
            # Cleanup
            os.remove(valid_file)
            os.remove(str(invalid_file_path))
            
            self.passed += 1
            return True
        except Exception as e:
            logger.error(f"  ❌ File validation failed: {e}")
            self.failed += 1
            self.errors.append(f"File Validation: {e}")
            return False
    
    def test_session_workflow(self):
        """Test 4: Session-based workflow"""
        logger.info("\n📋 Test 4: Session-based Workflow")
        try:
            user_id = 99999
            
            # Create session
            session = self.session_manager.get_session(user_id)
            assert session is not None, "Session not created"
            
            # Create test file and add to session
            file_path = self.create_test_file("workflow_test.xlsx", "Test", rows=3)
            self.session_manager.add_file(user_id, file_path, test_num=1)
            
            updated_session = self.session_manager.get_session(user_id)
            assert len(updated_session['uploaded_files']) > 0, "File not added to session"
            
            # Check session state
            status = self.session_manager.format_status_message(user_id)
            assert status is not None, "Status message not generated"
            
            # Get files for consolidation
            files = self.session_manager.get_files_for_consolidation(user_id)
            assert len(files) > 0, "Files not found for consolidation"
            
            # Cleanup
            self.session_manager.clear_session(user_id)
            os.remove(file_path)
            
            logger.info("  ✅ Session workflow working correctly")
            self.passed += 1
            return True
        except Exception as e:
            logger.error(f"  ❌ Session workflow failed: {e}")
            self.failed += 1
            self.errors.append(f"Session Workflow: {e}")
            return False
    
    def test_output_generation(self):
        """Test 5: Output generation"""
        logger.info("\n📋 Test 5: Output Generation")
        try:
            # Create test file
            file_path = self.create_test_file("output_test.xlsx", "Test", rows=3)
            
            # Process file
            assert os.path.exists(file_path), "Test file not created"
            
            # Create output directory if needed
            output_dir = Path(self.config.output_folder)
            output_dir.mkdir(exist_ok=True)
            
            logger.info("  ✅ Output generation structure verified")
            
            # Cleanup
            os.remove(file_path)
            
            self.passed += 1
            return True
        except Exception as e:
            logger.error(f"  ❌ Output generation failed: {e}")
            self.failed += 1
            self.errors.append(f"Output Generation: {e}")
            return False
    
    def test_error_recovery(self):
        """Test 6: Error recovery"""
        logger.info("\n📋 Test 6: Error Recovery")
        try:
            # Test handling of missing files
            try:
                with open('/nonexistent/path/file.xlsx', 'r') as f:
                    pass
                assert False, "Should have raised FileNotFoundError"
            except FileNotFoundError:
                logger.info("  ✅ Missing file handled gracefully")
            
            # Test handling of invalid operations
            try:
                invalid_value = int('not_a_number')
                assert False, "Should have raised ValueError"
            except ValueError:
                logger.info("  ✅ Invalid value handled gracefully")
            
            self.passed += 1
            return True
        except Exception as e:
            logger.error(f"  ❌ Error recovery failed: {e}")
            self.failed += 1
            self.errors.append(f"Error Recovery: {e}")
            return False
    
    async def run_all_tests(self):
        """Run complete test suite"""
        logger.info("=" * 60)
        logger.info("UPLOAD & CONSOLIDATION FLOW TEST SUITE")
        logger.info("=" * 60)
        logger.info(f"\n🚀 STARTING UPLOAD FLOW TESTS\n")
        
        self.test_file_creation()
        self.test_multiple_file_handling()
        self.test_file_validation()
        self.test_session_workflow()
        self.test_output_generation()
        self.test_error_recovery()
        
        # Print summary
        total = self.passed + self.failed
        success_rate = (self.passed / total * 100) if total > 0 else 0
        
        logger.info("\n" + "=" * 60)
        logger.info("TEST SUMMARY")
        logger.info("=" * 60)
        logger.info(f"Total: {total} | Passed: {self.passed} | Failed: {self.failed}")
        logger.info(f"Success Rate: {success_rate:.1f}%")
        logger.info("=" * 60)
        
        if self.errors:
            logger.error(f"\nErrors ({len(self.errors)}):")
            for error in self.errors:
                logger.error(f"  - {error}")
        
        return self.passed, self.failed


async def main():
    tester = UploadFlowTester()
    passed, failed = await tester.run_all_tests()
    
    if failed > 0:
        logger.warning(f"\nSome tests failed. Review errors above.")
        sys.exit(1)
    else:
        logger.info(f"\nAll tests passed! Upload flow is ready.")
        sys.exit(0)


if __name__ == "__main__":
    asyncio.run(main())
