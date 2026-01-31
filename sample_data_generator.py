#!/usr/bin/env python3
"""
Sample Data Generator for MLJResultsCompiler
Generates realistic test data for testing and demonstration
"""

import pandas as pd
import numpy as np
import random
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Sample participant names and emails
FIRST_NAMES = [
    'Abdulhamid', 'Abdullahi', 'Abdulsalam', 'Abdurrahman', 'Abiodun',
    'Alice', 'Bob', 'Charlie', 'David', 'Emily', 'Frank', 'Grace',
    'Henry', 'Iris', 'John', 'Kelly', 'Lisa', 'Michael', 'Nancy', 'Oliver'
]

LAST_NAMES = [
    'Abubakar', 'Bala', 'Gambo', 'Smith', 'Brown', 'Johnson', 'Williams',
    'Davis', 'Wilson', 'Moore', 'Taylor', 'Anderson', 'Thomas', 'Jackson',
    'White', 'Harris', 'Martin', 'Thompson', 'Garcia', 'Martinez'
]

EMAIL_DOMAINS = ['gmail.com', 'yahoo.com', 'outlook.com', 'test.com', 'example.com']

def generate_participants(count: int, seed: int = 42) -> list:
    """Generate random participant data"""
    random.seed(seed)
    np.random.seed(seed)
    
    participants = []
    used_emails = set()
    
    for _ in range(count):
        first = random.choice(FIRST_NAMES)
        last = random.choice(LAST_NAMES)
        name = f"{first} {last}"
        
        # Generate unique email
        base_email = f"{first.lower()}{last.lower()}{random.randint(1, 999)}@{random.choice(EMAIL_DOMAINS)}"
        while base_email in used_emails:
            base_email = f"{first.lower()}{last.lower()}{random.randint(1, 999)}@{random.choice(EMAIL_DOMAINS)}"
        used_emails.add(base_email)
        
        # Generate score (0-100%)
        score = round(random.uniform(50, 100), 1)
        
        participants.append({
            'name': name,
            'email': base_email,
            'score': score
        })
    
    return participants

def create_test_file(
    filename: str,
    participants: list,
    test_num: int,
    question_count: int = 25,
    color_hex: str = None,
    missing_ratio: float = 0.0
):
    """
    Create a test XLSX file with realistic structure
    
    Args:
        filename: Output filename
        participants: List of participant dicts with 'name', 'email', 'score'
        test_num: Test number (1-5)
        question_count: Number of questions to simulate
        color_hex: Color to apply (e.g., '87CEEB' for sky blue)
        missing_ratio: Ratio of participants to exclude (0.0 to 1.0)
    """
    
    # Simulate missing participants
    if missing_ratio > 0:
        num_to_remove = int(len(participants) * missing_ratio)
        participants = participants[:-num_to_remove]
    
    # Create question columns (Q1, Q2, ... Q{question_count})
    question_cols = [f'Q{i+1}' for i in range(question_count)]
    
    # Column name variations by test (to simulate real data)
    if test_num == 1:
        name_col = ' Full Names'  # Leading space
    elif test_num == 2:
        name_col = 'Full names'  # lowercase
    elif test_num == 3:
        name_col = 'Full Names'
    elif test_num == 4:
        name_col = 'Full Names'
    else:  # test_num == 5
        name_col = 'Name'  # Different name!
    
    # Build data
    data = []
    for p in participants:
        row = {
            name_col: p['name'],
            'Email': p['email'],
            'Result': f"{p['score']}%",
        }
        
        # Add question answers (random True/False)
        for q in question_cols:
            row[q] = random.choice(['True', 'False'])
        
        data.append(row)
    
    # Create dataframe
    df = pd.DataFrame(data)
    
    # Save to XLSX
    df.to_excel(filename, index=False)
    
    # Apply color formatting if specified
    if color_hex:
        wb = load_workbook(filename)
        ws = wb.active
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
        
        # Apply to all data rows (skip header)
        for row in ws.iter_rows(min_row=2, max_row=len(data) + 1):
            for cell in row:
                cell.fill = fill
        
        wb.save(filename)
    
    print(f"✓ Created {filename} ({len(data)} participants)")

def generate_test_suite():
    """Generate 5 test files with realistic structure"""
    output_folder = Path('input')
    output_folder.mkdir(exist_ok=True)
    
    print("\nGenerating sample test data...\n")
    
    # Color assignments
    colors = {
        1: None,        # White (no color)
        2: '87CEEB',    # Sky Blue
        3: 'FFFF00',    # Yellow
        4: '556B2F',    # Army Green
        5: 'FF0000',    # Red
    }
    
    # Question counts vary
    question_counts = {
        1: 26,
        2: 25,
        3: 40,
        4: 30,
        5: 19,
    }
    
    # Participant counts vary
    participant_counts = {
        1: 89,
        2: 92,
        3: 85,
        4: 85,
        5: 86,
    }
    
    # Generate base participants for Test 1
    base_participants = generate_participants(participant_counts[1])
    
    # Create Test 1
    create_test_file(
        output_folder / f'TEST_1_Obstetrics_Gynecology_JANUARY_2026.xlsx',
        base_participants,
        test_num=1,
        question_count=question_counts[1],
        color_hex=colors[1]
    )
    
    # Create Tests 2-5 with partial overlap
    for test_num in [2, 3, 4, 5]:
        # Keep 70-80% from Test 1, add 20-30% new participants
        overlap_ratio = random.uniform(0.70, 0.80)
        overlap_count = int(len(base_participants) * overlap_ratio)
        
        # Take some from Test 1
        test_participants = base_participants[:overlap_count]
        
        # Add new participants
        new_count = participant_counts[test_num] - overlap_count
        new_participants = generate_participants(new_count, seed=test_num * 100)
        test_participants.extend(new_participants)
        
        # Randomize order
        random.shuffle(test_participants)
        
        create_test_file(
            output_folder / f'TEST_{test_num}_Obstetrics_Gynecology_JANUARY_2026.xlsx',
            test_participants,
            test_num=test_num,
            question_count=question_counts[test_num],
            color_hex=colors[test_num]
        )
    
    print("\n✓ All test files generated in input/ folder")
    print("\nYou can now run: python results_compiler_bot.py")

if __name__ == '__main__':
    generate_test_suite()
