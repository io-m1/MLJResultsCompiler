"""
Color Configuration for Test Results
Maps each test number to its corresponding color code for Excel formatting
"""

from openpyxl.styles import PatternFill

# Base color definitions for each test
BASE_COLORS = {
    1: {'name': 'White', 'rgb': 'FFFFFF'},
    2: {'name': 'Sky Blue', 'rgb': '87CEEB'},
    3: {'name': 'Yellow', 'rgb': 'FFFF00'},
    4: {'name': 'Army Green', 'rgb': '556B2F'},
    5: {'name': 'Red', 'rgb': 'FF0000'},
    6: {'name': 'Light Green', 'rgb': '90EE90'},
    7: {'name': 'Light Coral', 'rgb': 'F08080'},
    8: {'name': 'Khaki', 'rgb': 'F0E68C'},
    9: {'name': 'Plum', 'rgb': 'DDA0DD'},
    10: {'name': 'Cyan', 'rgb': '00FFFF'},
}

# Extended color palette for tests beyond 10
EXTENDED_PALETTE = [
    'B0C4DE', '20B2AA', 'FFDAB9', 'E6E6FA', 'FFB6C1',
    'DEB887', '8FBC8F', 'FFFFE0', 'FFA07A', 'D3D3D3'
]

# Build TEST_COLORS dynamically
TEST_COLORS = {}
for i in range(1, 11):
    rgb = BASE_COLORS[i]['rgb']
    TEST_COLORS[i] = {
        'name': BASE_COLORS[i]['name'],
        'rgb': rgb,
        'fill': PatternFill(start_color=rgb, end_color=rgb, fill_type='solid')
    }

def get_fill_for_test(test_number: int) -> PatternFill:
    """
    Get the PatternFill object for a given test number (supports unlimited tests)
    
    Args:
        test_number (int): Test number (1+)
        
    Returns:
        PatternFill: The color fill object for that test
    """
    if test_number < 1:
        raise ValueError(f"Invalid test number: {test_number}. Must be >= 1")
    
    # If in base colors, use predefined
    if test_number in TEST_COLORS:
        return TEST_COLORS[test_number]['fill']
    
    # For tests beyond 10, cycle through extended palette
    palette_idx = (test_number - 11) % len(EXTENDED_PALETTE)
    rgb = EXTENDED_PALETTE[palette_idx]
    return PatternFill(start_color=rgb, end_color=rgb, fill_type='solid')


def get_color_name(test_number):
    """Get the name of the color for a test"""
    if test_number not in TEST_COLORS:
        raise ValueError(f"Invalid test number: {test_number}. Must be 1-5")
    return TEST_COLORS[test_number]['name']
