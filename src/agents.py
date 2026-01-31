"""
Agent-based architecture for MLJ Results Compiler
Implements Think-Act-Observe pattern with specialized agents
"""

import logging
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import openpyxl

logger = logging.getLogger(__name__)


class ValidationAgent:
    """Agent for validating uploaded Excel files"""
    
    REQUIRED_COLUMNS = ['Full Name', 'Email', 'Score', 'Result']
    VALID_EXTENSIONS = ['.xlsx']
    
    @staticmethod
    def think(files: List[str]) -> Dict:
        """Think: Analyze file structure and content"""
        analysis = {
            'valid_files': [],
            'invalid_files': [],
            'issues': [],
            'file_count': len(files)
        }
        
        for file_path in files:
            result = ValidationAgent._validate_single_file(file_path)
            if result['valid']:
                analysis['valid_files'].append(file_path)
            else:
                analysis['invalid_files'].append(file_path)
                analysis['issues'].extend(result['issues'])
        
        return analysis
    
    @staticmethod
    def _validate_single_file(file_path: str) -> Dict:
        """Validate a single Excel file"""
        try:
            path = Path(file_path)
            
            # Check extension
            if path.suffix.lower() not in ValidationAgent.VALID_EXTENSIONS:
                return {
                    'valid': False,
                    'issues': [f"Invalid file extension: {path.suffix}. Use .xlsx"]
                }
            
            # Load and check structure
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            if not ws:
                return {
                    'valid': False,
                    'issues': [f"{path.name}: No worksheets found"]
                }
            
            # Get headers
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            if not headers:
                return {
                    'valid': False,
                    'issues': [f"{path.name}: No header row found"]
                }
            
            # Check for required columns
            missing_cols = []
            for req_col in ValidationAgent.REQUIRED_COLUMNS:
                if req_col not in headers:
                    missing_cols.append(req_col)
            
            if missing_cols:
                return {
                    'valid': False,
                    'issues': [f"{path.name}: Missing columns: {', '.join(missing_cols)}"]
                }
            
            # Check data rows
            row_count = ws.max_row - 1
            if row_count == 0:
                return {
                    'valid': False,
                    'issues': [f"{path.name}: No data rows found"]
                }
            
            wb.close()
            
            return {
                'valid': True,
                'rows': row_count,
                'columns': headers,
                'issues': []
            }
            
        except Exception as e:
            return {
                'valid': False,
                'issues': [f"{Path(file_path).name}: {str(e)}"]
            }
    
    @staticmethod
    def act(analysis: Dict) -> Tuple[bool, str]:
        """Act: Decide based on analysis"""
        if not analysis['valid_files']:
            message = "❌ No valid files found.\n\n"
            for issue in analysis['issues'][:3]:
                message += f"• {issue}\n"
            if len(analysis['issues']) > 3:
                message += f"• ... and {len(analysis['issues']) - 3} more issues"
            return False, message
        
        success_msg = f"✅ Validated {len(analysis['valid_files'])} file(s)"
        
        if analysis['invalid_files']:
            success_msg += f"\n⚠️ {len(analysis['invalid_files'])} file(s) had issues:\n"
            for issue in analysis['issues'][:2]:
                success_msg += f"• {issue}\n"
        
        return True, success_msg
    
    @staticmethod
    def observe(is_valid: bool, analysis: Dict) -> Dict:
        """Observe: Return insights for next steps"""
        return {
            'is_valid': is_valid,
            'files_to_process': analysis['valid_files'],
            'file_count': len(analysis['valid_files']),
            'issues': analysis['issues']
        }


class ProcessingAgent:
    """Agent for processing and consolidating data"""
    
    @staticmethod
    def think(files: List[str], processor) -> Dict:
        """Think: Plan processing steps"""
        return {
            'files': files,
            'file_count': len(files),
            'processor': processor,
            'step': 'loading'
        }
    
    @staticmethod
    def act(plan: Dict) -> Tuple[bool, Dict, str]:
        """Act: Execute processing"""
        try:
            processor = plan['processor']
            
            # Load all tests
            loaded = processor.load_all_tests(max_tests=5)
            
            if loaded == 0:
                return False, {}, "❌ Failed to load files"
            
            # Consolidate
            consolidated = processor.consolidate_results()
            
            if not consolidated:
                return False, {}, "❌ Failed to consolidate results"
            
            return True, {
                'consolidated_data': consolidated,
                'participant_count': len(consolidated),
                'processor': processor
            }, f"✅ Processed {loaded} files, {len(consolidated)} participants"
            
        except Exception as e:
            logger.error(f"Processing error: {str(e)}")
            return False, {}, f"❌ Processing failed: {str(e)}"
    
    @staticmethod
    def observe(is_success: bool, result: Dict) -> Dict:
        """Observe: Ready for export"""
        return {
            'ready_to_export': is_success,
            'data': result.get('consolidated_data'),
            'participant_count': result.get('participant_count', 0),
            'processor': result.get('processor')
        }


class ExportAgent:
    """Agent for exporting in multiple formats"""
    
    FORMATS = {
        'xlsx': {'name': 'Excel', 'extension': '.xlsx'},
        'pdf': {'name': 'PDF', 'extension': '.pdf'},
        'docx': {'name': 'Word', 'extension': '.docx'}
    }
    
    @staticmethod
    def think(format_choice: str, data: Dict, processor, output_dir: str) -> Dict:
        """Think: Plan export steps"""
        return {
            'format': format_choice,
            'data': data,
            'processor': processor,
            'output_dir': output_dir,
            'format_info': ExportAgent.FORMATS.get(format_choice, {})
        }
    
    @staticmethod
    def act(plan: Dict) -> Tuple[bool, Path, str]:
        """Act: Execute export"""
        try:
            format_choice = plan['format']
            processor = plan['processor']
            output_dir = Path(plan['output_dir'])
            data = plan['data']
            
            output_dir.mkdir(parents=True, exist_ok=True)
            
            if format_choice == 'xlsx':
                filename = 'Consolidated_Results.xlsx'
                output_file = output_dir / filename
                processor.save_consolidated_file(data, filename)
                
            elif format_choice == 'pdf':
                filename = 'Consolidated_Results.pdf'
                output_file = output_dir / filename
                processor.save_as_pdf(data, filename)
                
            elif format_choice == 'docx':
                filename = 'Consolidated_Results.docx'
                output_file = output_dir / filename
                processor.save_as_docx(data, filename)
            else:
                return False, None, f"❌ Unknown format: {format_choice}"
            
            if not output_file.exists():
                return False, None, f"❌ Export file not created"
            
            format_name = plan['format_info'].get('name', 'Unknown')
            return True, output_file, f"✅ Exported as {format_name}"
            
        except Exception as e:
            logger.error(f"Export error: {str(e)}")
            return False, None, f"❌ Export failed: {str(e)}"
    
    @staticmethod
    def observe(is_success: bool, file_path: Optional[Path]) -> Dict:
        """Observe: File ready for delivery"""
        return {
            'export_success': is_success,
            'file_path': file_path,
            'file_size': file_path.stat().st_size if file_path and file_path.exists() else 0
        }


class ErrorRecoveryAgent:
    """Agent for handling errors and suggesting fixes"""
    
    ERROR_SOLUTIONS = {
        'missing_columns': "Please ensure your Excel files have these columns: Full Name, Email, Score, Result",
        'no_data': "Your Excel files appear to be empty. Please add data rows.",
        'invalid_format': "Please use .xlsx (Excel) file format only.",
        'processing_failed': "There was an issue processing your files. Please try again.",
        'export_failed': "Could not save the file. Please try a different format."
    }
    
    @staticmethod
    def think(error: str, user_context: Dict) -> Dict:
        """Think: Analyze error and context"""
        return {
            'error': error,
            'user_id': user_context.get('user_id'),
            'retry_count': user_context.get('retry_count', 0),
            'previous_attempts': user_context.get('attempts', [])
        }
    
    @staticmethod
    def act(analysis: Dict) -> Dict:
        """Act: Determine recovery strategy"""
        error = analysis['error']
        retry_count = analysis['retry_count']
        
        # Determine if we should retry
        should_retry = retry_count < 2
        
        # Find relevant solution
        solution = "Please check your files and try again."
        for key, msg in ErrorRecoveryAgent.ERROR_SOLUTIONS.items():
            if key.lower() in error.lower():
                solution = msg
                break
        
        return {
            'can_recover': should_retry,
            'solution': solution,
            'retry_count': retry_count + 1,
            'user_guidance': f"🔄 Attempt {retry_count + 1}/3: {solution}"
        }
    
    @staticmethod
    def observe(recovery: Dict) -> Dict:
        """Observe: Recovery plan ready"""
        return {
            'recovery_possible': recovery['can_recover'],
            'user_message': recovery['user_guidance'],
            'solution': recovery['solution'],
            'next_step': 'retry' if recovery['can_recover'] else 'help'
        }


class ContextMemoryAgent:
    """Agent for tracking user preferences and history"""
    
    @staticmethod
    def initialize_context(user_id: int) -> Dict:
        """Initialize context for new user"""
        return {
            'user_id': user_id,
            'preferred_format': None,
            'past_exports': [],
            'file_patterns': [],
            'retry_count': 0,
            'total_files_processed': 0,
            'successful_exports': 0
        }
    
    @staticmethod
    def record_export(context: Dict, format_choice: str, success: bool):
        """Record successful export"""
        context['past_exports'].append({
            'format': format_choice,
            'success': success,
            'timestamp': __import__('datetime').datetime.now().isoformat()
        })
        
        if success:
            if not context['preferred_format']:
                context['preferred_format'] = format_choice
            context['successful_exports'] += 1
    
    @staticmethod
    def get_user_preference(context: Dict) -> Optional[str]:
        """Get user's preferred format based on history"""
        return context.get('preferred_format')
    
    @staticmethod
    def get_summary(context: Dict) -> str:
        """Get user statistics"""
        return (
            f"📊 Your Stats:\n"
            f"• Successful exports: {context['successful_exports']}\n"
            f"• Preferred format: {context['preferred_format'] or 'Not set yet'}\n"
            f"• Total exports: {len(context['past_exports'])}"
        )
