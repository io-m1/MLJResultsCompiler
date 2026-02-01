# -*- coding: utf-8 -*-
"""
Data Integrity Verification Tests
Validates that consolidation preserves all data and headers are correct
"""

import logging
import sys
from pathlib import Path
from typing import Dict, List, Tuple

logger = logging.getLogger(__name__)


class DataIntegrityValidator:
    """
    Validates data consolidation for:
    - No duplicate email loss
    - Proper headers
    - Data completeness
    - Column detection
    """
    
    def __init__(self):
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self.info: List[str] = []
    
    def validate_consolidation(self, 
                             test_data_before: Dict,
                             consolidated_data: Dict) -> Dict:
        """
        Validate that consolidation preserved all data.
        
        Args:
            test_data_before: Original data {test_num: {email: data}}
            consolidated_data: Consolidated {email: data}
            
        Returns:
            Validation result with issues/warnings
        """
        self.errors = []
        self.warnings = []
        self.info = []
        
        # 1. Check for unique emails in source
        all_unique_emails = set()
        for test_num, test_data in test_data_before.items():
            all_unique_emails.update(test_data.keys())
        
        self.info.append(f"✓ Source data: {len(all_unique_emails)} unique emails across all tests")
        
        # 2. Check consolidated has all emails
        consolidated_emails = set(consolidated_data.keys())
        
        if len(consolidated_emails) != len(all_unique_emails):
            missing = all_unique_emails - consolidated_emails
            self.errors.append(
                f"❌ DATA LOSS: {len(missing)} emails missing from consolidation\n"
                f"   Source: {len(all_unique_emails)} unique emails\n"
                f"   Output: {len(consolidated_emails)} emails\n"
                f"   Missing: {missing}"
            )
        else:
            self.info.append(f"✓ All {len(all_unique_emails)} emails present in consolidated data")
        
        # 3. Check for duplicate handling
        duplicates_by_test = self._find_duplicates_in_test_data(test_data_before)
        if duplicates_by_test:
            for test_num, dups in duplicates_by_test.items():
                self.warnings.append(
                    f"⚠ Test {test_num} has {len(dups)} duplicate emails (same person tested twice)\n"
                    f"   These are merged by email (keeping all scores)"
                )
        
        # 4. Validate column structure
        if consolidated_data:
            first_entry = next(iter(consolidated_data.values()))
            expected_keys = {'name'}
            actual_keys = set(first_entry.keys())
            
            # Check for test score columns
            test_score_columns = [k for k in actual_keys if k.startswith('test_') and k.endswith('_score')]
            if test_score_columns:
                self.info.append(f"✓ Found {len(test_score_columns)} test score columns: {sorted(test_score_columns)}")
            else:
                self.errors.append("❌ No test score columns found in consolidated data")
        
        # 5. Validate headers will be correct
        self._validate_headers(consolidated_data)
        
        return {
            "valid": len(self.errors) == 0,
            "errors": self.errors,
            "warnings": self.warnings,
            "info": self.info,
            "stats": {
                "source_unique_emails": len(all_unique_emails),
                "consolidated_emails": len(consolidated_emails),
                "data_loss_percent": (1 - len(consolidated_emails) / len(all_unique_emails) * 100) if all_unique_emails else 0
            }
        }
    
    def _find_duplicates_in_test_data(self, test_data: Dict) -> Dict:
        """Find duplicate emails across tests"""
        duplicates = {}
        all_emails_by_test = {}
        
        for test_num, data in test_data.items():
            all_emails_by_test[test_num] = set(data.keys())
        
        # Find emails that appear in multiple tests
        for test_num1, emails1 in all_emails_by_test.items():
            for test_num2, emails2 in all_emails_by_test.items():
                if test_num1 < test_num2:
                    overlaps = emails1 & emails2
                    if overlaps:
                        key = f"Test_{test_num1}_vs_Test_{test_num2}"
                        if key not in duplicates:
                            duplicates[key] = list(overlaps)
        
        return duplicates
    
    def _validate_headers(self, consolidated_data: Dict):
        """Validate that headers will be correct (not using participant data)"""
        if not consolidated_data:
            return
        
        first_entry = next(iter(consolidated_data.values()))
        
        # Check if name/email look like real participant data (good)
        first_name = first_entry.get('name')
        if first_name and len(first_name) > 50:
            self.errors.append(
                f"❌ HEADER ERROR: First entry has suspicious name length: {len(first_name)} chars\n"
                f"   This might indicate participant data is being used as headers"
            )
        else:
            self.info.append(f"✓ Header structure looks correct (name field is reasonable)")
    
    def validate_excel_output(self, excel_path: Path) -> Dict:
        """
        Validate the actual Excel file structure.
        
        Args:
            excel_path: Path to Excel file
            
        Returns:
            Validation result
        """
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            
            # Get first row (headers)
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value))
            
            self.info.append(f"✓ Excel file has {len(headers)} columns")
            
            # Check if headers are correct
            if headers[0] == "Full Name" and headers[1] == "Email":
                self.info.append(f"✓ Headers are correct: {headers[:5]}...")
            else:
                self.errors.append(
                    f"❌ HEADER ERROR: Expected ['Full Name', 'Email', ...]\n"
                    f"   Got: {headers}"
                )
            
            # Count data rows
            data_rows = ws.max_row - 1  # Exclude header
            self.info.append(f"✓ Excel has {data_rows} data rows")
            
            # Check for empty cells in key columns
            empty_names = 0
            empty_emails = 0
            for row in range(2, ws.max_row + 1):
                if not ws[f'A{row}'].value:
                    empty_names += 1
                if not ws[f'B{row}'].value:
                    empty_emails += 1
            
            if empty_names > 0 or empty_emails > 0:
                self.warnings.append(
                    f"⚠ Found {empty_names} empty names, {empty_emails} empty emails in output"
                )
            else:
                self.info.append(f"✓ All {data_rows} rows have name and email")
            
            return {
                "valid": len(self.errors) == 0,
                "errors": self.errors,
                "warnings": self.warnings,
                "info": self.info,
                "file_stats": {
                    "headers": headers,
                    "data_rows": data_rows,
                    "total_columns": len(headers)
                }
            }
            
        except Exception as e:
            self.errors.append(f"❌ Failed to validate Excel file: {e}")
            return {
                "valid": False,
                "error": str(e)
            }
    
    def print_report(self, title: str = "Data Integrity Report"):
        """Print validation report"""
        print(f"\n{'='*60}")
        print(f"  {title}")
        print(f"{'='*60}\n")
        
        if self.info:
            print("[INFO] INFO:")
            for msg in self.info:
                print(f"  {msg}")
            print()
        
        if self.warnings:
            print("⚠️  WARNINGS:")
            for msg in self.warnings:
                print(f"  {msg}")
            print()
        
        if self.errors:
            print("❌ ERRORS:")
            for msg in self.errors:
                print(f"  {msg}")
            print()
        
        if not self.errors:
            print("[OK] VALIDATION PASSED - No critical issues found!\n")


# Singleton
_validator = None

def get_validator() -> DataIntegrityValidator:
    """Get or create validator instance"""
    global _validator
    if _validator is None:
        _validator = DataIntegrityValidator()
    return _validator
