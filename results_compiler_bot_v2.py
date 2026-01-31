#!/usr/bin/env python3
"""
MLJResultsCompiler - Scalable Results Compilation Bot (Enhanced)
Consolidates ANY number of test result sheets with agentic capabilities
Author: MLJ Results Compiler
Date: January 31, 2026
Version: 2.0 - Scalable & Agentic
"""

import os
import sys
import logging
import pandas as pd
import numpy as np
from pathlib import Path
from typing import Tuple, List, Optional, Dict
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import glob

# Import configuration and agents
sys.path.insert(0, str(Path(__file__).parent))
from config import BotConfig, ProcessingStrategy, MergeStrategy, DataValidationLevel
from agents import AgentOrchestrator, AgentReport

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('compiler_execution.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


class ScalableResultsCompiler:
    """
    Enhanced compiler that handles ANY number of test files
    Includes agentic decision-making and autonomous optimization
    """
    
    def __init__(self, config: Optional[BotConfig] = None, input_folder: str = 'input', output_folder: str = 'output'):
        """Initialize the scalable compiler with configuration"""
        # Use provided config or defaults
        if config:
            self.config = config
            if input_folder != 'input':
                self.config.input_folder = input_folder
            if output_folder != 'output':
                self.config.output_folder = output_folder
        else:
            self.config = BotConfig(input_folder=input_folder, output_folder=output_folder)
        
        self.input_folder = Path(self.config.input_folder)
        self.output_folder = Path(self.config.output_folder)
        self.test_files = {}        # Dict[str, Path] - test names to file paths
        self.dataframes = {}        # Dict[str, pd.DataFrame] - test data
        self.merged_df = None       # Final merged dataframe
        self.file_order = []        # Track file processing order
        
        # Initialize agentic system
        self.orchestrator = AgentOrchestrator(config=self.config) if self.config.enable_agents else None
        self.agent_reports = []     # Store all agent reports
        
        # Create output folder if it doesn't exist
        self.output_folder.mkdir(exist_ok=True)
        
        logger.info(f"ScalableResultsCompiler v2.0 initialized")
        logger.info(f"Input folder: {self.input_folder.absolute()}")
        logger.info(f"Output folder: {self.output_folder.absolute()}")
        logger.info(f"Processing strategy: {self.config.processing_strategy.value}")
        logger.info(f"Agents enabled: {self.config.enable_agents}")
    
    def find_test_files(self) -> bool:
        """
        Dynamically find ALL test files matching the pattern
        Works with any number of files (1, 5, 50, 1000+)
        """
        logger.info(f"Searching for test files matching pattern: {self.config.file_pattern}")
        
        if not self.input_folder.exists():
            logger.error(f"Input folder not found: {self.input_folder}")
            return False
        
        # Build glob pattern
        pattern = self.config.file_pattern
        if not self.config.case_sensitive_pattern:
            # Case-insensitive search
            pattern = pattern.replace('.xlsx', '') + '*.xlsx'
            matches = list(self.input_folder.glob(pattern))
        else:
            matches = list(self.input_folder.glob(pattern))
        
        if not matches:
            logger.warning("No files found matching pattern")
            return False
        
        # Sort matches for consistent ordering
        matches = sorted(matches)
        
        # Extract test names from filenames
        for filepath in matches:
            filename = filepath.stem  # filename without extension
            
            # Try to extract test number/name
            test_name = self._extract_test_name(filename)
            if not test_name:
                test_name = filename  # Use filename as fallback
            
            self.test_files[test_name] = filepath
            self.file_order.append(test_name)
        
        logger.info(f"✓ Found {len(self.test_files)} test files:")
        for test_name, filepath in sorted(self.test_files.items()):
            logger.info(f"  - {test_name}: {filepath.name}")
        
        # Check minimum files requirement
        if len(self.test_files) < self.config.min_files_required:
            logger.error(f"Found {len(self.test_files)} files, but {self.config.min_files_required} required")
            return False
        
        # Warn if over maximum
        if len(self.test_files) > self.config.max_files_allowed:
            logger.warning(f"Found {len(self.test_files)} files, limiting to {self.config.max_files_allowed}")
            self.test_files = dict(list(self.test_files.items())[:self.config.max_files_allowed])
        
        return True
    
    @staticmethod
    def _extract_test_name(filename: str) -> Optional[str]:
        """Extract test name/number from filename"""
        # Try to find TEST_1, TEST_2, etc.
        match = re.search(r'[Tt][Ee][Ss][Tt][\s_]?(\d+)', filename)
        if match:
            return f"Test {match.group(1)}"
        
        # Try to find just a number
        match = re.search(r'^(\d+)', filename)
        if match:
            return f"Test {match.group(1)}"
        
        return None
    
    def detect_column_names(self, df: pd.DataFrame) -> Tuple[Optional[str], Optional[str], Optional[str]]:
        """
        Intelligently detect column names using configuration mapping
        Handles variations defined in column_mapping
        """
        columns = df.columns.tolist()
        
        # Normalize column names for matching
        columns_lower = [col.lower().strip() for col in columns]
        
        # Detect Name column
        name_col = None
        for i, col_lower in enumerate(columns_lower):
            if col_lower in self.config.column_mapping.name_variations:
                name_col = columns[i]
                break
        
        # Detect Email column
        email_col = None
        for i, col_lower in enumerate(columns_lower):
            if col_lower in self.config.column_mapping.email_variations:
                email_col = columns[i]
                break
        
        # Detect Score column
        score_col = None
        for i, col_lower in enumerate(columns_lower):
            if col_lower in self.config.column_mapping.score_variations:
                score_col = columns[i]
                break
        
        if not all([name_col, email_col, score_col]):
            logger.warning(f"Could not detect all columns. Found: Name={name_col}, Email={email_col}, Score={score_col}")
        
        return name_col, email_col, score_col
    
    def load_test_file(self, test_name: str, filepath: Path) -> bool:
        """Load and extract data from a single test file"""
        logger.info(f"Loading {test_name} from {filepath.name}...")
        
        try:
            # Read XLSX file
            df = pd.read_excel(filepath)
            logger.debug(f"{test_name}: Loaded {len(df)} rows, {len(df.columns)} columns")
            
            # Validate with agents
            if self.config.validation_agent_enabled and self.orchestrator:
                val_report = self.orchestrator.validation_agent.validate_dataframe(df, test_name)
                self.agent_reports.append(val_report)
                if val_report.issues_found > 0:
                    logger.warning(f"{test_name}: {val_report.message}")
            
            # Detect column names
            name_col, email_col, score_col = self.detect_column_names(df)
            
            if not all([name_col, email_col, score_col]):
                logger.error(f"{test_name}: Could not detect all required columns")
                return False
            
            # Extract only required columns
            df_extracted = df[[name_col, email_col, score_col]].copy()
            
            # Rename columns to standard names
            df_extracted.columns = ['Full Name', 'Email', f'{test_name} Score']
            
            # Clean data
            df_extracted['Full Name'] = df_extracted['Full Name'].astype(str).str.strip()
            df_extracted['Email'] = df_extracted['Email'].astype(str).str.strip().str.lower()
            
            # Parse score
            df_extracted[f'{test_name} Score'] = df_extracted[f'{test_name} Score'].apply(self._parse_score)
            
            # Remove rows with missing critical data
            df_extracted = df_extracted.dropna(subset=['Full Name', 'Email'])
            df_extracted = df_extracted[df_extracted['Email'] != '']
            
            logger.info(f"✓ {test_name}: Extracted {len(df_extracted)} participants")
            self.dataframes[test_name] = df_extracted
            return True
            
        except Exception as e:
            logger.error(f"Error loading {test_name}: {str(e)}")
            return False
    
    def load_all_test_files(self) -> bool:
        """Load all discovered test files"""
        logger.info(f"Loading {len(self.test_files)} test files...")
        
        success_count = 0
        for test_name in self.file_order:
            if self.load_test_file(test_name, self.test_files[test_name]):
                success_count += 1
        
        logger.info(f"Successfully loaded {success_count} of {len(self.test_files)} test files")
        
        # Check minimum requirement
        if success_count < self.config.min_files_required:
            logger.error(f"Insufficient files loaded: {success_count} < {self.config.min_files_required}")
            return False
        
        return True
    
    def merge_tests(self) -> bool:
        """Merge all test dataframes using configured merge strategy"""
        logger.info(f"Merging {len(self.dataframes)} tests using {self.config.merge_strategy.value} strategy...")
        
        if not self.dataframes:
            logger.error("No test dataframes loaded")
            return False
        
        try:
            # Get first file as base
            first_test = self.file_order[0]
            if first_test not in self.dataframes:
                logger.error(f"First test file {first_test} not loaded")
                return False
            
            self.merged_df = self.dataframes[first_test].copy()
            logger.debug(f"Starting merge with {first_test}: {len(self.merged_df)} rows")
            
            # Merge with remaining tests
            for test_name in self.file_order[1:]:
                if test_name in self.dataframes:
                    if self.config.merge_strategy == MergeStrategy.EMAIL:
                        self.merged_df = self._merge_on_email(self.merged_df, self.dataframes[test_name], test_name)
                    else:
                        logger.warning(f"Merge strategy {self.config.merge_strategy.value} not fully implemented, using EMAIL")
                        self.merged_df = self._merge_on_email(self.merged_df, self.dataframes[test_name], test_name)
                    
                    logger.debug(f"After merging {test_name}: {len(self.merged_df)} rows")
            
            logger.info(f"✓ Merge complete: {len(self.merged_df)} unique participants")
            return True
            
        except Exception as e:
            logger.error(f"Error during merge: {str(e)}")
            return False
    
    @staticmethod
    def _merge_on_email(left_df: pd.DataFrame, right_df: pd.DataFrame, test_name: str) -> pd.DataFrame:
        """Merge two dataframes on email (outer join)"""
        merged = pd.merge(
            left_df,
            right_df,
            on='Email',
            how='outer',
            suffixes=('', f'_new')
        )
        
        # Handle name column conflicts - keep from left if available
        if 'Full Name_new' in merged.columns:
            merged['Full Name'] = merged['Full Name'].fillna(merged['Full Name_new'])
            merged.drop('Full Name_new', axis=1, inplace=True)
        
        return merged
    
    def clean_and_sort(self) -> bool:
        """Remove duplicates and sort based on configuration"""
        logger.info("Cleaning and sorting results...")
        
        try:
            initial_count = len(self.merged_df)
            
            # Remove duplicate emails
            if self.config.remove_duplicates:
                self.merged_df = self.merged_df.drop_duplicates(subset=['Email'], keep='first')
                duplicates_removed = initial_count - len(self.merged_df)
                if duplicates_removed > 0:
                    logger.info(f"Removed {duplicates_removed} duplicate emails")
            
            # Remove empty rows
            if self.config.remove_empty_rows:
                before = len(self.merged_df)
                self.merged_df = self.merged_df.dropna(how='all')
                logger.info(f"Removed {before - len(self.merged_df)} empty rows")
            
            # Sort
            if self.config.sort_by == 'name':
                self.merged_df = self.merged_df.sort_values(
                    'Full Name',
                    key=lambda x: x.str.lower(),
                    ignore_index=True
                )
            elif self.config.sort_by == 'email':
                self.merged_df = self.merged_df.sort_values(
                    'Email',
                    ignore_index=True
                )
            
            logger.info(f"✓ Cleaned and sorted {len(self.merged_df)} participants")
            return True
            
        except Exception as e:
            logger.error(f"Error during cleanup: {str(e)}")
            return False
    
    def format_scores(self) -> bool:
        """Format scores as percentages"""
        logger.info("Formatting scores...")
        
        try:
            score_columns = [col for col in self.merged_df.columns if 'Score' in col]
            
            for col in score_columns:
                self.merged_df[col] = self.merged_df[col].apply(
                    lambda x: f"{x:.1f}%" if pd.notna(x) else ""
                )
            
            logger.info(f"✓ Formatted {len(score_columns)} score columns")
            return True
            
        except Exception as e:
            logger.error(f"Error formatting scores: {str(e)}")
            return False
    
    def export_to_xlsx(self, filename: str = 'Consolidated_Results.xlsx') -> bool:
        """Export to professionally formatted XLSX"""
        logger.info(f"Exporting to {filename}...")
        
        try:
            output_path = self.output_folder / filename
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Results'
            
            # Define styles
            header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            header_font = Font(bold=True, size=12)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Write header row
            headers = self.merged_df.columns.tolist()
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Write data rows with intelligent coloring
            color_map = self.config.colors.colors
            for row_idx, (_, row) in enumerate(self.merged_df.iterrows(), 2):
                for col_idx, (col_name, value) in enumerate(zip(headers, row), 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    
                    # Apply color based on test column
                    for test_key, color in color_map.items():
                        if test_key.replace('_', ' ') in col_name:
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                            break
                    
                    # Center alignment for scores
                    if 'Score' in col_name:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for col_idx, header in enumerate(headers, 1):
                max_length = len(str(header))
                for row in self.merged_df.itertuples(index=False):
                    max_length = max(max_length, len(str(row[col_idx-1])))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"✓ Exported to {output_path}")
            logger.info(f"✓ File contains {len(self.merged_df)} participants and {len(headers)} columns")
            
            return True
            
        except Exception as e:
            logger.error(f"Error exporting to XLSX: {str(e)}")
            return False
    
    def generate_insights(self) -> Dict:
        """Generate data insights using quality agent"""
        if not self.config.quality_agent_enabled or not self.orchestrator:
            return {}
        
        logger.info("Generating insights...")
        insights = self.orchestrator.generate_insights(self.merged_df)
        logger.info(f"Data Quality Score: {insights.get('completeness_percent', 'N/A')}%")
        return insights
    
    def generate_report(self):
        """Generate and print comprehensive summary report"""
        logger.info("\n" + "="*60)
        logger.info("COMPILATION SUMMARY")
        logger.info("="*60)
        
        logger.info(f"Files processed: {len(self.test_files)}")
        logger.info(f"Total unique participants: {len(self.merged_df) if self.merged_df is not None else 0}")
        
        if self.merged_df is not None:
            # Count participants per test
            for test_name in self.file_order:
                score_col = f'{test_name} Score'
                if score_col in self.merged_df.columns:
                    non_empty = (self.merged_df[score_col] != '').sum()
                    logger.info(f"{test_name}: {non_empty} participants")
        
        logger.info(f"\nOutput file: {self.output_folder / 'Consolidated_Results.xlsx'}")
        logger.info(f"Log file: compiler_execution.log")
        
        # Print agent reports
        if self.config.enable_agents and self.agent_reports:
            logger.info("\n" + "-"*60)
            logger.info("AGENT REPORTS")
            logger.info("-"*60)
            for report in self.agent_reports:
                logger.info(str(report))
        
        logger.info("="*60 + "\n")
    
    def run(self) -> bool:
        """Execute the full compilation pipeline"""
        logger.info("Starting MLJ Results Compilation Bot v2.0 (Scalable & Agentic)...")
        logger.info(f"Timestamp: {datetime.now()}")
        
        start_time = datetime.now()
        
        try:
            # Step 1: Find test files
            if not self.find_test_files():
                logger.error("Could not find test files")
                return False
            
            # Step 2: Load test files
            if not self.load_all_test_files():
                logger.error("Could not load test files")
                return False
            
            # Step 3: Merge tests
            if not self.merge_tests():
                logger.error("Could not merge test results")
                return False
            
            # Step 4: Clean and sort
            if not self.clean_and_sort():
                logger.error("Could not clean and sort results")
                return False
            
            # Step 5: Format scores
            if not self.format_scores():
                logger.error("Could not format scores")
                return False
            
            # Step 6: Export to XLSX
            if not self.export_to_xlsx():
                logger.error("Could not export results")
                return False
            
            # Step 7: Generate insights
            self.generate_insights()
            
            # Generate report
            self.generate_report()
            
            end_time = datetime.now()
            duration = (end_time - start_time).total_seconds()
            
            logger.info(f"✓ COMPILATION COMPLETE in {duration:.2f} seconds")
            return True
            
        except Exception as e:
            logger.error(f"Unexpected error: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            return False
    
    @staticmethod
    def _parse_score(score) -> float:
        """Parse score value"""
        if pd.isna(score) or score == '':
            return np.nan if not isinstance(score, str) else None
        
        score_str = str(score).strip().replace('%', '')
        try:
            return float(score_str)
        except ValueError:
            return np.nan


def main():
    """Main entry point"""
    # Parse command line arguments
    input_folder = sys.argv[1] if len(sys.argv) > 1 else 'input'
    output_folder = sys.argv[2] if len(sys.argv) > 2 else 'output'
    
    # Create compiler and run
    compiler = ScalableResultsCompiler(input_folder=input_folder, output_folder=output_folder)
    
    if compiler.run():
        sys.exit(0)
    else:
        sys.exit(1)


if __name__ == '__main__':
    main()
