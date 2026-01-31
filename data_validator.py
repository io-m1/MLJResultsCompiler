#!/usr/bin/env python3
"""
Quality Assurance Validator for Test Results
Pre- and post-processing validation to catch errors before they affect final results
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os
import json
from typing import Dict, List, Tuple
from datetime import datetime

class TestDataValidator:
    def __init__(self):
        self.validation_report = {
            'timestamp': datetime.now().isoformat(),
            'pre_processing_checks': {},
            'post_processing_checks': {},
            'issues_found': [],
            'warnings': [],
            'overall_status': 'PASS'
        }
    
    def validate_input_files(self, input_dir: str) -> Dict:
        """Validate all input test files before processing"""
        print("\n[PRE-PROCESSING] Validating input files...")
        
        report = {
            'files_found': [],
            'files_missing': [],
            'sheet_structure': {},
            'data_quality': {}
        }
        
        expected_tests = {1, 2, 3, 4, 5}
        found_tests = set()
        
        for file in os.listdir(input_dir):
            if not file.endswith('.xlsx'):
                continue
            
            filepath = os.path.join(input_dir, file)
            
            # Determine test number
            test_num = None
            for t in expected_tests:
                if f'TEST_{t}' in file or f'test_{t}' in file:
                    test_num = t
                    found_tests.add(t)
                    break
            
            if test_num is None:
                if 'ultrasonography' in file.lower() and 'test_5' in file.lower():
                    test_num = 5
                    found_tests.add(5)
            
            if test_num is None:
                continue
            
            file_info = {
                'file': file,
                'test': test_num,
                'filepath': filepath,
                'validation': {}
            }
            
            try:
                # Load and validate structure
                wb = load_workbook(filepath)
                
                # Check for Responses sheet
                if 'Responses' not in wb.sheetnames:
                    file_info['validation']['sheet_error'] = \
                        f"'Responses' sheet not found. Available: {wb.sheetnames}"
                    self.validation_report['issues_found'].append(
                        f"TEST_{test_num}: Missing 'Responses' sheet"
                    )
                else:
                    ws = wb['Responses']
                    file_info['validation']['rows'] = ws.max_row
                    file_info['validation']['columns'] = ws.max_column
                    
                    # Check required columns
                    headers = [cell.value for cell in ws[1]]
                    required_patterns = ['Name', 'Email', 'Result']
                    
                    found_cols = {
                        'name': False,
                        'email': False,
                        'result': False
                    }
                    
                    for header in headers:
                        if header:
                            h_lower = str(header).lower()
                            if 'name' in h_lower:
                                found_cols['name'] = True
                            if 'email' in h_lower:
                                found_cols['email'] = True
                            if 'result' in h_lower or 'mark' in h_lower:
                                found_cols['result'] = True
                    
                    missing_cols = [k for k, v in found_cols.items() if not v]
                    
                    if missing_cols:
                        file_info['validation']['column_warning'] = \
                            f"Missing expected columns: {missing_cols}"
                        self.validation_report['warnings'].append(
                            f"TEST_{test_num}: Missing columns: {missing_cols}"
                        )
                    
                    # Data quality checks
                    df = pd.read_excel(filepath, sheet_name='Responses')
                    
                    data_quality = {
                        'total_rows': len(df),
                        'empty_rows': df.isna().all(axis=1).sum(),
                        'null_counts': df.isna().sum().to_dict()
                    }
                    
                    if data_quality['empty_rows'] > 0:
                        self.validation_report['warnings'].append(
                            f"TEST_{test_num}: {data_quality['empty_rows']} empty rows"
                        )
                    
                    file_info['validation']['data_quality'] = data_quality
                    file_info['validation']['status'] = 'OK' if not missing_cols else 'WARNING'
                    
                report['files_found'].append(file_info)
                
            except Exception as e:
                file_info['validation']['error'] = str(e)
                file_info['validation']['status'] = 'ERROR'
                self.validation_report['issues_found'].append(
                    f"TEST_{test_num} ({file}): {str(e)}"
                )
                report['files_found'].append(file_info)
        
        # Check for missing tests
        missing = expected_tests - found_tests
        if missing:
            report['files_missing'] = list(missing)
            self.validation_report['warnings'].append(
                f"Missing test files: {missing}"
            )
        
        report['status'] = 'OK' if not missing else 'WARNING'
        self.validation_report['pre_processing_checks'] = report
        
        self._print_validation_report(report)
        return report
    
    def validate_output_file(self, output_filepath: str) -> Dict:
        """Validate final output file"""
        print("\n[POST-PROCESSING] Validating output file...")
        
        report = {
            'file': output_filepath,
            'exists': os.path.exists(output_filepath),
            'structure': {},
            'data_integrity': {},
            'formula_checks': {}
        }
        
        if not report['exists']:
            self.validation_report['issues_found'].append(
                "Output file does not exist"
            )
            report['status'] = 'ERROR'
            return report
        
        try:
            # Check file size
            file_size = os.path.getsize(output_filepath)
            report['file_size_kb'] = round(file_size / 1024, 2)
            
            if file_size < 10000:  # Less than 10KB seems too small
                self.validation_report['warnings'].append(
                    f"Output file seems small: {report['file_size_kb']}KB"
                )
            
            # Load and validate structure
            wb = load_workbook(output_filepath)
            
            if 'Responses' not in wb.sheetnames:
                raise ValueError("'Responses' sheet not found in output")
            
            ws = wb['Responses']
            
            # Check headers
            expected_headers = ['S/N', 'NAMES', 'EMAIL', 'TEST 1', 'TEST 2', 
                               'TEST 3', 'TEST 4', 'TEST 5', 'GRP DISCUSSION', 
                               'TOTAL MARK', 'SCORE', 'STATUS']
            
            actual_headers = [cell.value for cell in ws[1]]
            
            header_match = actual_headers == expected_headers
            report['structure']['headers_match'] = header_match
            
            if not header_match:
                self.validation_report['warnings'].append(
                    f"Headers don't match expected. Expected: {expected_headers}, "
                    f"Got: {actual_headers}"
                )
            
            # Data integrity checks
            data_rows = ws.max_row - 1
            report['structure']['data_rows'] = data_rows
            
            # Check for completeness
            issues = []
            for row_num in range(2, min(ws.max_row + 1, 10)):  # Check first 9 rows
                # Check required columns filled
                name = ws[f'B{row_num}'].value
                email = ws[f'C{row_num}'].value
                
                if not name:
                    issues.append(f"Row {row_num}: Missing name")
                if not email:
                    issues.append(f"Row {row_num}: Missing email")
            
            if issues:
                report['data_integrity']['issues'] = issues[:5]  # First 5
                self.validation_report['warnings'].extend(issues)
            
            # Formula checks
            formula_issues = []
            for row_num in range(2, min(ws.max_row + 1, 10)):
                total_cell = ws[f'J{row_num}']
                score_cell = ws[f'K{row_num}']
                status_cell = ws[f'L{row_num}']
                
                # Check formulas exist
                if not total_cell.value or not isinstance(total_cell.value, str) or '=' not in str(total_cell.value):
                    formula_issues.append(f"Row {row_num}: TOTAL MARK missing formula")
                
                if not score_cell.value or not isinstance(score_cell.value, str) or '=' not in str(score_cell.value):
                    formula_issues.append(f"Row {row_num}: SCORE missing formula")
                
                if not status_cell.value or not isinstance(status_cell.value, str) or '=' not in str(status_cell.value):
                    formula_issues.append(f"Row {row_num}: STATUS missing formula")
            
            if formula_issues:
                report['formula_checks']['issues'] = formula_issues
                self.validation_report['issues_found'].extend(formula_issues)
            else:
                report['formula_checks']['status'] = 'OK'
            
            report['status'] = 'OK' if not issues and not formula_issues else 'WARNING'
            
        except Exception as e:
            report['error'] = str(e)
            report['status'] = 'ERROR'
            self.validation_report['issues_found'].append(str(e))
        
        self.validation_report['post_processing_checks'] = report
        self._print_output_validation(report)
        return report
    
    def check_result_counts(self, output_filepath: str, input_dir: str) -> Dict:
        """Cross-check participant counts"""
        print("\n[CONSISTENCY] Checking participant counts...")
        
        report = {
            'by_test': {},
            'final_count': 0,
            'coverage_analysis': {}
        }
        
        try:
            # Count participants in each input test
            for test_num in range(1, 6):
                for file in os.listdir(input_dir):
                    if f'TEST_{test_num}' not in file and not (test_num == 5 and 'ultrasonography' in file.lower()):
                        continue
                    
                    filepath = os.path.join(input_dir, file)
                    if not file.endswith('.xlsx'):
                        continue
                    
                    try:
                        df = pd.read_excel(filepath, sheet_name='Responses')
                        count = len(df[df.iloc[:, 2].notna()])  # Count rows with name
                        report['by_test'][f'TEST_{test_num}'] = count
                    except:
                        pass
            
            # Count final output
            if os.path.exists(output_filepath):
                df_output = pd.read_excel(output_filepath, sheet_name='Responses')
                report['final_count'] = len(df_output) - 1  # Exclude header
                
                # Coverage analysis
                for col in ['TEST 1', 'TEST 2', 'TEST 3', 'TEST 4', 'TEST 5']:
                    non_null = df_output[col].notna().sum()
                    coverage = round((non_null / report['final_count'] * 100), 1) if report['final_count'] > 0 else 0
                    report['coverage_analysis'][col] = {
                        'participants': non_null,
                        'coverage_percent': coverage
                    }
            
            report['status'] = 'OK'
            
        except Exception as e:
            report['error'] = str(e)
            report['status'] = 'ERROR'
        
        self._print_count_analysis(report)
        return report
    
    def _print_validation_report(self, report):
        """Pretty print validation report"""
        print("\n  File Validation Summary:")
        print(f"    Files found: {len(report['files_found'])}")
        print(f"    Files missing: {report.get('files_missing', [])}")
        
        for file_info in report['files_found']:
            status = file_info['validation'].get('status', '?')
            rows = file_info['validation'].get('rows', '?')
            print(f"    TEST_{file_info['test']}: {status} ({rows} rows)")
            
            if 'column_warning' in file_info['validation']:
                print(f"      ⚠️  {file_info['validation']['column_warning']}")
            
            if 'error' in file_info['validation']:
                print(f"      ❌ {file_info['validation']['error']}")
    
    def _print_output_validation(self, report):
        """Pretty print output validation"""
        print("\n  Output File Validation:")
        print(f"    File: {os.path.basename(report.get('file', 'unknown'))}")
        print(f"    Status: {report.get('status', '?')}")
        print(f"    Exists: {report.get('exists', False)}")
        print(f"    Size: {report.get('file_size_kb', 0)} KB")
        print(f"    Data rows: {report['structure'].get('data_rows', '?')}")
        
        if report['structure'].get('headers_match') is False:
            print("    ⚠️  Headers don't match expected")
        
        if report['formula_checks'].get('issues'):
            for issue in report['formula_checks']['issues'][:3]:
                print(f"    ❌ {issue}")
    
    def _print_count_analysis(self, report):
        """Pretty print count analysis"""
        print("\n  Participant Count Analysis:")
        for test, count in report['by_test'].items():
            print(f"    {test}: {count} participants")
        
        print(f"    Final output: {report['final_count']} participants")
        
        if report['coverage_analysis']:
            print("\n  Coverage by Test:")
            for test, analysis in report['coverage_analysis'].items():
                coverage = analysis['coverage_percent']
                print(f"    {test}: {analysis['participants']}/{report['final_count']} ({coverage}%)")
    
    def save_report(self, output_dir: str) -> str:
        """Save validation report as JSON"""
        if self.validation_report['issues_found'] or self.validation_report['warnings']:
            self.validation_report['overall_status'] = 'WARNING' if self.validation_report['warnings'] and not self.validation_report['issues_found'] else 'FAIL'
        
        filename = f"validation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        filepath = os.path.join(output_dir, filename)
        
        os.makedirs(output_dir, exist_ok=True)
        
        with open(filepath, 'w') as f:
            json.dump(self.validation_report, f, indent=2)
        
        print(f"\n[SAVED] Validation report: {filepath}")
        return filepath
    
    def run_full_validation(self, input_dir: str, output_filepath: str = None) -> bool:
        """Run complete validation suite"""
        print("\n" + "="*70)
        print("TEST RESULTS VALIDATION SUITE")
        print("="*70)
        
        # Pre-processing checks
        self.validate_input_files(input_dir)
        
        # Post-processing checks (if output provided)
        if output_filepath and os.path.exists(output_filepath):
            self.validate_output_file(output_filepath)
            self.check_result_counts(output_filepath, input_dir)
        
        # Overall status
        has_errors = len(self.validation_report['issues_found']) > 0
        has_warnings = len(self.validation_report['warnings']) > 0
        
        print("\n" + "="*70)
        if has_errors:
            print(f"❌ VALIDATION FAILED - {len(self.validation_report['issues_found'])} errors found")
            self.validation_report['overall_status'] = 'FAIL'
        elif has_warnings:
            print(f"⚠️  VALIDATION PASSED WITH WARNINGS - {len(self.validation_report['warnings'])} warnings")
            self.validation_report['overall_status'] = 'WARNING'
        else:
            print("✅ VALIDATION PASSED - All checks OK")
            self.validation_report['overall_status'] = 'PASS'
        print("="*70 + "\n")
        
        return not has_errors


if __name__ == '__main__':
    import sys
    
    if len(sys.argv) < 2:
        print("Usage: python data_validator.py <input_dir> [output_file]")
        sys.exit(1)
    
    input_dir = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    validator = TestDataValidator()
    success = validator.run_full_validation(input_dir, output_file)
    validator.save_report(os.path.dirname(output_file) if output_file else input_dir)
    
    sys.exit(0 if success else 1)
